import os
import sys
import xml.etree.ElementTree as ET
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def parse_junit(junit_path: str):
    """
    Reads JUnit XML and returns:
      - rows: list of dicts with test case details
      - summary: totals dict
    """
    tree = ET.parse(junit_path)
    root = tree.getroot()

    # JUnit can be <testsuites> containing <testsuite>, or a single <testsuite>
    suites = []
    if root.tag == "testsuites":
        suites = root.findall("testsuite")
    elif root.tag == "testsuite":
        suites = [root]
    else:
        raise ValueError(f"Unexpected JUnit root tag: {root.tag}")

    rows = []
    total = passed = failed = skipped = 0

    for suite in suites:
        suite_name = suite.attrib.get("name", "")
        for tc in suite.findall("testcase"):
            total += 1
            name = tc.attrib.get("name", "")
            classname = tc.attrib.get("classname", "")
            time_s = tc.attrib.get("time", "")

            # Determine result
            if tc.find("failure") is not None or tc.find("error") is not None:
                result = "FAIL"
                failed += 1
                fail_node = tc.find("failure") or tc.find("error")
                message = (fail_node.attrib.get("message", "") if fail_node is not None else "")
            elif tc.find("skipped") is not None:
                result = "SKIP"
                skipped += 1
                message = ""
            else:
                result = "PASS"
                passed += 1
                message = ""

            rows.append({
                "Result": result,
                "Test Name": name,
                "Class": classname,
                "Suite": suite_name,
                "Duration(s)": time_s,
                "Message": message
            })

    summary = {
        "Total": total,
        "Pass": passed,
        "Fail": failed,
        "Skip": skipped
    }
    return rows, summary


def autosize_columns(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def main():
    junit_path = os.environ.get("JUNIT_PATH", "test-results/junit.xml")
    out_dir = os.environ.get("OUT_DIR", "test-results")
    out_file = os.environ.get("OUT_FILE", "playwright-results.xlsx")

    if not os.path.exists(junit_path):
        print(f"ERROR: JUnit file not found at: {junit_path}")
        sys.exit(1)

    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, out_file)

    rows, summary = parse_junit(junit_path)

    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Generated (UTC)", datetime.now(timezone.utc).isoformat()])
    ws_summary.append([])
    ws_summary.append(["Metric", "Value"])
    for k, v in summary.items():
        ws_summary.append([k, v])
    autosize_columns(ws_summary)

    # Sheet 2: Results
    ws = wb.create_sheet("Results")
    headers = ["Result", "Test Name", "Class", "Suite", "Duration(s)", "Message"]
    ws.append(headers)

    for r in rows:
        ws.append([r[h] for h in headers])

    autosize_columns(ws)

    wb.save(out_path)
    print(f"✅ Wrote Excel results: {out_path}")


if __name__ == "__main__":
    main()
